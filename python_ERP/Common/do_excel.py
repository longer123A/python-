#!/usr/bin/python3
# -*- coding: utf-8 -*-
#Author: xu
from openpyxl import load_workbook
from python_ERP.Common.dir_config import *
class DoExcel:

    def __init__(self,file_path,sheet_name):
        self.file_path=file_path
        self.sheet_name=sheet_name

    def do_excel(self,key,row,col):
        """
        :param key: 字典key值
        :param row: 行
        :param col: 列
        :return: 以字典形式返回
        """
        wb=load_workbook(self.file_path)
                   #打开excel工作簿
        sheet=wb[self.sheet_name]
                   #获取页面
        sub_data={}                                             #用sub_data字典装载数据
        sub_data[key]=sheet.cell(row,col).value
        return sub_data

    def write_back(self,row,col,new_value):
                  #响应结果返回函数
        """
                  :param row: 行
                  :param col: 列
                  :param new_value:新值
                  """
        wb=load_workbook(self.file_path)
        sheet=wb[self.sheet_name]
        sheet.cell(row,col).value=new_value
        wb.save(self.file_path)



if __name__ == '__main__':
    s=DoExcel(testcases_dir,'aaa').do_excel('key',1,1)
    print(s['key'])
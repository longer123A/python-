from openpyxl import load_workbook
from post_res.common.pro_path5 import *
class DoExcel:

    def __init__(self,file_path,sheet_name):
        self.file_path=file_path
        self.sheet_name=sheet_name


    def do_excel(self):
        wb=load_workbook(self.file_path)
                   #打开excel工作簿
        sheet=wb[self.sheet_name]
                   #获取页面

        tel=self.get_tel()
                  #获取初始化手机号码
        self.update_tel(tel+1)
                  #更新初始化手机号码

        test_data=[]
        for i in range(2,sheet.max_row+1):
            sub_data={}                                             #用sub_data字典装载数据
            sub_data['CaseId']=sheet.cell(i,1).value                #获取第2行第1列的数据
            sub_data['Tltle']=sheet.cell(i,2).value
            sub_data['Url']=sheet.cell(i,3).value
            sub_data['Param']=sheet.cell(i,4).value
            if sub_data['Param'].find('${tel}')!=-1:
                                #判断'${tel}'是否存在
                sub_data['Param']=str(sub_data['Param'])
                                #把sub_data['Param']转化为字符串
                sub_data['Param']=sub_data['Param'].replace('${tel}',str(tel))
                                #把数据中的'${tel}'替换

            sub_data['Expected']=sheet.cell(i,5).value
            sub_data['Method']=sheet.cell(i,6).value
            test_data.append(sub_data)                              #添加至test_data列表中

        return test_data

    def write_back(self,row,col,new_value):
                  #响应结果返回函数
        wb=load_workbook(self.file_path)
        sheet=wb[self.sheet_name]
        sheet.cell(row,col).value=new_value
        wb.save(self.file_path)

    def get_tel(self):
                   #获取Excel里面的初始化和数
        wb=load_workbook(self.file_path)
        sheet=wb['Sheet2']
        tel=sheet.cell(1,1).value
        return tel

    def update_tel(self,new_tel):
                #更新初始化手机号码的函数
        wb=load_workbook(self.file_path)
        sheet=wb['Sheet2']
        sheet.cell(1,1).value=new_tel
        wb.save(self.file_path)

if __name__ == '__main__':
    data=DoExcel(test_data_path,'Sheet1').do_excel()
    print(data)
